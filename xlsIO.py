from datetime import datetime, time

from openpyxl.cell.cell import Cell
from scheduleclasses import Schedule, TestItem
from typing import Any, List, Optional, Union
from dataclasses import dataclass

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from PySide6.QtCore import QObject, Signal, Slot

tick = b'\xe2\x9c\x93'.decode('utf-8')


class XlsIO(QObject):
    """A class for handling file IO and parsing and handling xls files"""

    schedule_to_update = Signal(Schedule)

    def __init__(self) -> None:
        super().__init__()
        self.__file_path = ""
        self.__worksheet = None
        self.parser = Xlparser()

    @property
    def filepath(self) -> Any:
        return self.__file_path

    @filepath.setter
    def filepath(self, path: str) -> None:
        self.__file_path = path

    @property
    def xl_worksheet(self) -> Worksheet:
        return self.__worksheet

    @xl_worksheet.setter
    def xl_worksheet(self, sheet: Worksheet) -> None:
        self.__worksheet = sheet

    def xlsheet_from_path(self, path: str) -> Worksheet:
        self.xl_worksheet = openpyxl.load_workbook(
            path,
            data_only=True).active

    @Slot(str)
    def import_excel(self, import_path: str) -> None:
        self.filepath = import_path
        self.xlsheet_from_path(self.filepath)
        self.parser.parse_xl(self.xl_worksheet)
        self.schedule_to_update.emit(
            self.parser.schedule
        )


@dataclass
class Xlparser:
    def __init__(self) -> None:
        self.header_row_bottom: Optional[int] = None
        self.merged_cells: list = []
        self.indexes: dict = {}
        self.schedule: Schedule = Schedule()
        self.xldata: Optional[Worksheet] = None

    def find_merged_cells(self) -> None:
        self.merged_cells = self.xldata.merged_cells

    def get_header_info(self) -> None:
        self.header_row_bottom = None
        for row in self.xldata.rows:
            for cell in row:
                cell: Cell
                if type(cell.value) is not str:
                    pass
                elif "SFI" in cell.value:
                    self.indexes.update(sfi=cell.column)
                elif "Test Item" in cell.value:
                    self.indexes.update(item_name=cell.column)
                elif "Class" in cell.value:
                    self.header_row_bottom = cell.row
                    self.indexes.update({'class': cell.column})
                elif "Flag" in cell.value:
                    self.indexes.update(flag=cell.column)
                elif "Owner" in cell.value:
                    self.indexes.update(owner=cell.column)
                elif "To Be Recorded/Followed" in cell.value:
                    self.indexes.update(record_stat=cell.column)
                elif "Responsible" in cell.value:
                    self.indexes.update(responsible=cell.column)
                elif "Date" in cell.value:
                    self.indexes.update(date=cell.column)
                elif "Start Time" in cell.value:
                    self.indexes.update(start_time=cell.column)
                elif "Estimated" in cell.value:
                    self.indexes.update(
                        est=cell.column)
                elif "Cemre NB" in cell.value:
                    self.get_hull_and_owner(cell)

            if self.header_row_bottom is not None:
                break

    def get_cell_range(self, cell):
        for r in self.merged_cells:
            if cell.coordinate in r:
                return r

    def get_hull_and_owner(self, cell: Cell):
        for val in cell.value.split(' '):
            if 'NB' in val:
                self.schedule.hull_number = val
                owner_cell = self.xldata.cell(
                    cell.row + 1, cell.column
                )
                self.schedule.owner_firm = owner_cell.value.title()

    def add_rows2schedule(self, row: tuple) -> None:
        item = TestItem()

        for cell in row:
            if cell.coordinate in self.merged_cells:
                r = self.get_cell_range(cell)
                i_row, i_col = r.top[0]
                cell = self.xldata.cell(i_row, i_col)

            if cell.column == self.indexes['sfi']:
                item.sfi =\
                    cell.value if cell.value is not None else ''

            elif cell.column == self.indexes['item_name']:
                item.item_name =\
                    cell.value if cell.value is not None else ''

            elif cell.column == self.indexes['class']:
                item.class_attendance =\
                    Xlparser.correct_attendance(cell.value)\
                    if cell.value is not None else ''

            elif cell.column == self.indexes['flag']:
                item.flag_attendance =\
                    Xlparser.correct_attendance(cell.value)\
                    if cell.value is not None else ''

            elif cell.column == self.indexes['owner']:
                item.owner_attendance =\
                    Xlparser.correct_attendance(cell.value)\
                    if cell.value is not None else ''

            elif cell.column == self.indexes['record_stat']:
                item.record_status =\
                    cell.value if cell.value is not None else ''

            elif cell.column == self.indexes['responsible']:
                item.responsible_dept =\
                    cell.value if cell.value is not None else ''

            elif cell.column == self.indexes['date']:
                item.date =\
                    cell.value if cell.value is not None else ''

            elif cell.column == self.indexes['start_time']:
                item.start_hour =\
                    cell.value if cell.value is not None else ''

            elif cell.column == self.indexes['est']:
                item.est = Xlparser.correct_time_format(cell.value)

        if item.date == '' or\
                item.start_hour == '':
            return None
        else:
            self.schedule.add_item(item)

    @staticmethod
    def correct_attendance(char: str) -> str:
        if char == 'ü':
            return tick
        else:
            return char

    @staticmethod
    def correct_time_format(data: Union[datetime, time]) -> Union[time, str]:
        if isinstance(data, datetime):
            return time(data.hour, data.minute)
        if isinstance(data, time):
            return data
        else:
            return '-'

    def get_attendee_selection(self) -> List[str]:
        selection = []
        for row_num in range(self.header_row_bottom+1,
                             self.xldata.max_row):
            val = self.xldata.cell(
                    row_num, self.indexes['responsible']
                ).value
            if val not in selection and val is not None:
                selection.append(val)
        return selection

    def get_item_rows(self) -> None:
        min_row = self.header_row_bottom+1
        for row in self.xldata.iter_rows(min_row):
            self.add_rows2schedule(row)

    def parse_xl(self, xldata: Worksheet) -> None:
        self.xldata = xldata
        self.schedule.reset_items()
        self.get_header_info()
        self.schedule.responsible_selection = self.get_attendee_selection()
        self.find_merged_cells()
        self.get_item_rows()
