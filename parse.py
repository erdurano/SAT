import openpyxl
import datetime as dt


def parse_duration_str(xlduration_str):
    dur_pieces = xlduration_str.split(" ")

    if 'min' in dur_pieces[-1]:
        return dt.timedelta(minutes=int(dur_pieces[-2]))

    if 'hour' in dur_pieces[-1]:
        if '.' in dur_pieces[-2]:
            nums = dur_pieces[-2].split('.')
            return dt.timedelta(hours=int(nums[0]), minutes=int(nums[1]))
        else:
            return dt.timedelta(hours=int(dur_pieces[-2]))

    if 'day' in dur_pieces[-1]:
        return dt.timedelta(days=int(dur_pieces[-2]))


def parse_times(date=None, start_hour=None, duration=None):
    xldate = date
    xlhour = start_hour
    xlduration_str = duration

    if (xldate and xlhour):
        xlhour_delta = dt.timedelta(hours=xlhour.hour, minutes=xlhour.minute)
        start_time_dt = xldate + xlhour_delta
    else:
        start_time_dt = None

    if xlduration_str is None:
        end_time_dt = None
    else:
        end_time_dt = start_time_dt + parse_duration_str(xlduration_str)

    return {
        'start_datetime': start_time_dt,
        'estimated_finish': end_time_dt
    }


def parse_SAT_doc(file_path):

    def get_merged_cell_val(row, column):
        cell = row[column]
        if cell.coordinate not in merged_cells:
            return cell

        else:
            while cell.value is None:
                cell = worksheet.cell(row=cell.row-1, column=cell.column)
            return cell

    worksheet = openpyxl.load_workbook(file_path).active
    merged_cells = worksheet.merged_cells
    header_row = None
    test_agenda = []
    for row in worksheet.iter_rows():
        if header_row is None:
            for cell in row:
                if 'Class' == cell.value:
                    header_row = cell.row
                    class_column = cell.column

        elif header_row is not None:
            test_row = {
                "sfi": row[class_column - 3].value
                if row[class_column - 3].value is not None else "",
                "item_name": row[class_column - 2].value,
                "class_att": row[class_column - 1].value,
                "flag_att": row[class_column].value,
                "owner_att": row[class_column + 1].value,
                "rec_stat": row[class_column + 2].value,
                "resp_dept": row[class_column + 3].value,
            }

            time_frame = parse_times(
                get_merged_cell_val(row, class_column + 4).value,
                get_merged_cell_val(row, class_column + 5).value,
                get_merged_cell_val(row, class_column + 6).value,
            )
            test_row.update(time_frame)
            if (test_row["item_name"] is not None) and \
               (test_row["start_datetime"] is not None):
                test_agenda.append(test_row)

    return test_agenda


if __name__ == '__main__':
    filename = './example/5_NB67_SAT_Schedule_Rev_3.xlsx'
    result = parse_SAT_doc(filename)
    print(result)
    print(len(result))
